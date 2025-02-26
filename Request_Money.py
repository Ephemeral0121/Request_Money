import sys
import os
import json
from copy import copy
from PyQt5.QtWidgets import QFileDialog, QDesktopWidget, QMessageBox, QScrollArea, QSpinBox, QDateEdit, QApplication, QHBoxLayout, QWidget, QVBoxLayout, QLabel, QLineEdit, QPushButton, QComboBox, QRadioButton, QButtonGroup, QGroupBox, QFormLayout
from openpyxl import load_workbook
from datetime import datetime
from PyQt5.QtCore import QDate, Qt
from PyQt5.QtGui import QPalette, QIcon, QFont
from openpyxl.utils.cell import range_boundaries
from openpyxl.styles import Font, Alignment
from PyQt5.QtWidgets import QDialog, QFormLayout, QDialogButtonBox
from comtypes.client import CreateObject

#Config 클래스
class ConfigDialog(QDialog):

    def resource_path(self, relative_path):
        try:
            base_path = sys._MEIPASS
        except Exception:
            base_path = os.path.abspath(".")
        return os.path.join(base_path, relative_path)
    

    def __init__(self):
        super().__init__()
        self.bankAccountsLayouts = []  # 은행 계좌 정보를 저장할 리스트
        self.initUI()
        self.loadConfig()

    def initUI(self):
        self.setWindowTitle('회사 설정')
        self.setGeometry(100, 100, 800, 400)
        self.setWindowIcon(QIcon(self.resource_path("icon.png")))
        layout = QVBoxLayout()
        layout.setSpacing(10)  # 위젯 간의 간격 설정

        # 폰트 설정
        font = QFont("Arial", 10)
        self.setFont(font)

        formLayout = QFormLayout()
        formLayout.setFieldGrowthPolicy(QFormLayout.AllNonFixedFieldsGrow)
        self.companyNameEdit = QLineEdit()
        self.addressEdit = QLineEdit()
        self.documentNumberEdit = QLineEdit()
        self.accountHolderEdit = QLineEdit()

        formLayout.addRow('회사명:', self.companyNameEdit)
        formLayout.addRow('주소:', self.addressEdit)
        formLayout.addRow('문서번호:', self.documentNumberEdit)
        formLayout.addRow('예금주명:', self.accountHolderEdit)

        layout.addLayout(formLayout)

        # 스크롤 영역 스타일링
        scrollArea = QScrollArea()
        scrollArea.setWidgetResizable(True)
        scrollArea.setStyleSheet("QScrollArea { border: none; }")
        self.accountsWidget = QWidget()
        self.accountsSection = QVBoxLayout(self.accountsWidget)
        scrollArea.setWidget(self.accountsWidget)

        layout.addWidget(scrollArea)

        # 버튼 스타일링
        self.addAccountButton = QPushButton(QIcon(self.resource_path("add_icon.png")), '계좌 추가')
        self.addAccountButton.clicked.connect(lambda _: self.addAccountField())
        layout.addWidget(self.addAccountButton)

        buttonBox = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        buttonBox.accepted.connect(self.accept)
        buttonBox.rejected.connect(self.reject)
        layout.addWidget(buttonBox)

        self.setLayout(layout)
        self.center()


    def addAccountField(self, bankName='', accountNumber=''):
        accountLayout = QHBoxLayout()
        bankNameEdit = QLineEdit()
        bankNameEdit.setPlaceholderText('은행명')
        bankNameEdit.setText(bankName)
        accountNumberEdit = QLineEdit()
        accountNumberEdit.setPlaceholderText('계좌번호')
        accountNumberEdit.setText(accountNumber)

        # 삭제 버튼 추가 및 삭제 이벤트 핸들러 설정
        deleteButton = QPushButton('삭제')
        deleteButton.clicked.connect(lambda: self.deleteAccountField(accountLayout, bankNameEdit, accountNumberEdit, deleteButton))

        accountLayout.addWidget(bankNameEdit)
        accountLayout.addWidget(accountNumberEdit)
        accountLayout.addWidget(deleteButton)

        self.accountsSection.addLayout(accountLayout)
        self.bankAccountsLayouts.append((bankNameEdit, accountNumberEdit, accountLayout, deleteButton))

    def deleteAccountField(self, accountLayout, bankNameEdit, accountNumberEdit, deleteButton):
        bankNameEdit.deleteLater()
        accountNumberEdit.deleteLater()
        deleteButton.deleteLater()
        self.accountsSection.removeItem(accountLayout)
        self.bankAccountsLayouts.remove((bankNameEdit, accountNumberEdit, accountLayout, deleteButton))


    def loadConfig(self):
        try:
            with open('config.json', 'r') as configFile:
                config = json.load(configFile)
                self.companyNameEdit.setText(config.get('companyName', ''))
                self.addressEdit.setText(config.get('address', ''))
                self.documentNumberEdit.setText(config.get('documentNumber', ''))
                self.accountHolderEdit.setText(config.get('accountHolder', ''))

                # 저장된 계좌 정보를 사용하여 계좌 정보 입력란 생성
                for account in config.get('bankAccounts', []):
                    self.addAccountField(account.get('bankName', ''), account.get('accountNumber', ''))
        except FileNotFoundError:
            pass  # 파일이 없으면 아무것도 하지 않음


    def center(self):
        qr = self.frameGeometry()
        cp = QDesktopWidget().availableGeometry().center()
        qr.moveCenter(cp)
        self.move(qr.topLeft())


class Request_Money(QWidget):

    def resource_path(self, relative_path):
        try:
            base_path = sys._MEIPASS
        except Exception:
            base_path = os.path.abspath(".")
        return os.path.join(base_path, relative_path)
    
    def __init__(self):
        super().__init__()
        self.initUI()

    def initUI(self):
        mainLayout = QVBoxLayout()
        mainLayout.setSpacing(10)  # 위젯 간의 간격 조정
        mainLayout.setContentsMargins(10, 10, 10, 10)  # 레이아웃의 여백 조정
        self.setWindowIcon(QIcon(self.resource_path("icon.png")))

        # 회사 설정
        self.configButton = QPushButton('회사 설정')
        self.configButton.clicked.connect(self.openConfigDialog)
        mainLayout.addWidget(self.configButton)


        # 엔터티 유형 및 계산 유형 설정
        self.setupEntityTypeGroup(mainLayout)
        self.setupCalculationTypeGroup(mainLayout)

        # 수입금액 입력
        incomeLayout = QHBoxLayout()
        incomeLayout.addWidget(QLabel('기준금액:'))
        self.incomeEdit = QLineEdit()
        incomeLayout.addWidget(self.incomeEdit)
        mainLayout.addLayout(incomeLayout)

        # 원가계산 누진 및 기본계산 누진 설정
        self.setupCostProgression(mainLayout)
        self.setupBaseCalcProgression(mainLayout)

        # 인원수 설정
        self.setupNumPeople(mainLayout)

        # 결과 라벨 구성
        self.resultLabel = QLabel('기본보수:')
        self.resultLabel.setStyleSheet("font-weight: bold; color: green")

        self.additionalLabel = QLabel('')
        self.additionalLabel.setStyleSheet("font-weight: bold; color: green")

        self.final_resultLabel = QLabel('최종기본보수:')
        self.final_resultLabel.setStyleSheet("font-weight: bold; color: green")

        self.costProgressionLabel = QLabel('원가계산누진:')
        self.costProgressionLabel.setStyleSheet("font-weight: bold; color: blue")


        # 결과 라벨을 담을 스크롤 영역 설정
        self.scrollArea = QScrollArea()
        self.scrollArea.setWidgetResizable(True)
        self.scrollAreaWidgetContents = QWidget()
        self.scrollArea.setWidget(self.scrollAreaWidgetContents)
        self.resultLayout = QVBoxLayout(self.scrollAreaWidgetContents)
        self.scrollAreaWidgetContents.setLayout(self.resultLayout)

        # 기본보수, 추가보수, 최종기본보수 라벨을 스크롤 영역에 추가
        self.resultLayout.addWidget(self.resultLabel)
        self.resultLayout.addWidget(self.additionalLabel)
        self.resultLayout.addWidget(self.final_resultLabel)
        self.resultLayout.addWidget(self.costProgressionLabel)

        # 스크롤 영역을 메인 레이아웃에 추가
        mainLayout.addWidget(self.scrollArea)

        # 회사명 입력
        companyNameLayout = QHBoxLayout()
        companyNameLayout.addWidget(QLabel('거래처명:'))
        self.companyNameEdit = QLineEdit()
        companyNameLayout.addWidget(self.companyNameEdit)
        mainLayout.addLayout(companyNameLayout)

        # 납기일 설정
        self.setupDueDate(mainLayout)

        #기타감면 성실신고
        self.setupOtherReductionsAndRemuneration(mainLayout)

        # 엑셀 파일 생성 버튼
        self.createExcelButton = QPushButton('엑셀 파일 생성')
        mainLayout.addWidget(self.createExcelButton)

        # PDF 파일 생성 버튼
        self.createPdfButton = QPushButton('엑셀+PDF 파일 생성')
        self.createPdfButton.clicked.connect(self.create_pdf_file)
        mainLayout.addWidget(self.createPdfButton)

        # 모든 위젯 설정 후 이벤트 연결
        self.incomeEdit.textChanged.connect(self.calculateRemuneration)  # 입력 값 변경 시 계산 메소드 호출
        self.corporateButton.toggled.connect(self.calculateRemuneration)
        self.individualButton.toggled.connect(self.calculateRemuneration)
        self.typeAButton.toggled.connect(self.calculateRemuneration)
        self.typeBButton.toggled.connect(self.calculateRemuneration)
        self.createExcelButton.clicked.connect(self.create_excel_file)

        self.corporateButton.setChecked(True)
        self.typeBButton.setChecked(True)
        self.costProgressionGroup.buttons()[0].setChecked(True)
        self.baseCalcProgressionGroup.buttons()[0].setChecked(True)

        # 메시지 라벨
        self.messageLabel = QLabel()
        mainLayout.addWidget(self.messageLabel)

        self.applyStyleSheet()
        self.setGeometry(300, 300, 700, 1150)
        self.setLayout(mainLayout)
        self.setWindowTitle('조정보수청구')

    def applyStyleSheet(self):
        self.setStyleSheet("""
            QWidget {
                font-size: 14pt;
                font-family: 'Helvetica Neue', Helvetica, Arial, sans-serif;
            }

            QLineEdit, QDateEdit{
                border: 1px solid #ccc;
                border-radius: 8px;
                padding: 5px;
                background: #fff;
            }

            QPushButton {
                background-color: #007aff;
                color: #fff;
                border-radius: 8px;
                padding: 6px 12px;
                border: 1px solid #007aff;
                margin: 5px;
            }

            QPushButton:hover {
                background-color: #005ecb;
            }

            QRadioButton {
                spacing: 5px;
            }

            QScrollArea {
                border: 1px solid #ccc;
                border-radius: 8px;
            }

            QSpinBox:disabled, QSpinBox:!enabled {  # 비활성화된 QSpinBox 스타일
            background: #e6e6e6;
            color: #999;  # 글자색 변경
            }
        """)

    def openConfigDialog(self):
        self.configDialog = ConfigDialog()
        if self.configDialog.exec_():
            self.saveConfig()  # 인자 없이 메서드 호출


    def saveConfig(self):
        dialog = self.configDialog  # ConfigDialog 인스턴스를 self.configDialog에 저장해두었다고 가정
        config = {
            'companyName': dialog.companyNameEdit.text().strip(),
            'address': dialog.addressEdit.text().strip(),
            'documentNumber': dialog.documentNumberEdit.text().strip(),
            'accountHolder': dialog.accountHolderEdit.text().strip(),
            'bankAccounts': []  # 은행 계좌 정보를 저장할 리스트
        }

        # bankAccountsLayouts 리스트에서 필요한 두 개의 값만 언패킹합니다.
        for bankNameEdit, accountNumberEdit, _, _ in dialog.bankAccountsLayouts:
            bankAccount = {
                'bankName': bankNameEdit.text().strip(),
                'accountNumber': accountNumberEdit.text().strip()
            }
            config['bankAccounts'].append(bankAccount)

        with open('config.json', 'w') as configFile:
            json.dump(config, configFile, ensure_ascii=False, indent=4)

    def loadConfig(self):
        try:
            with open('config.json', 'r') as configFile:
                config = json.load(configFile)
                self.companyNameEdit.setText(config.get('companyName', ''))
                self.addressEdit.setText(config.get('address', ''))
                self.documentNumberEdit.setText(config.get('documentNumber', ''))
                self.accountHolderEdit.setText(config.get('accountHolder', ''))

                # 저장된 계좌 정보를 사용하여 계좌 정보 입력란 생성
                for account in config.get('bankAccounts', []):
                    self.addAccountField(account.get('bankName', ''), account.get('accountNumber', ''))
        except FileNotFoundError:
            pass  # 파일이 없으면 아무것도 하지 않음


    def setupEntityTypeGroup(self, mainLayout):
        entityTypeBox = QGroupBox("엔터티 유형 선택")
        entityTypeLayout = QHBoxLayout()
        entityTypeLayout.setSpacing(5)  # 라디오 버튼 사이의 간격 조정
        self.entityTypeGroup = QButtonGroup(self)
        self.corporateButton = QRadioButton('법인')
        self.individualButton = QRadioButton('개인')
        self.entityTypeGroup.addButton(self.corporateButton)
        self.entityTypeGroup.addButton(self.individualButton)
        entityTypeLayout.addWidget(self.corporateButton)
        entityTypeLayout.addWidget(self.individualButton)
        entityTypeBox.setLayout(entityTypeLayout)
        mainLayout.addWidget(entityTypeBox)


    def setupCalculationTypeGroup(self, mainLayout):
        calculationTypeBox = QGroupBox("계산 유형 선택")
        calculationTypeLayout = QHBoxLayout()
        self.calculationTypeGroup = QButtonGroup(self)
        self.typeAButton = QRadioButton('A유형')
        self.typeBButton = QRadioButton('B유형')
        self.calculationTypeGroup.addButton(self.typeAButton)
        self.calculationTypeGroup.addButton(self.typeBButton)
        self.typeAButton.toggled.connect(self.calculateRemuneration)
        self.typeBButton.toggled.connect(self.calculateRemuneration)
        calculationTypeLayout.addWidget(self.typeAButton)
        calculationTypeLayout.addWidget(self.typeBButton)
        calculationTypeBox.setLayout(calculationTypeLayout)
        self.typeBButton.setChecked(True)
        mainLayout.addWidget(calculationTypeBox)

    def setupCostProgression(self, mainLayout):
        costProgressionBox = QGroupBox("원가계산 누진 선택")
        costProgressionLayout = QHBoxLayout()
        self.costProgressionGroup = QButtonGroup(self)
        self.costProgressionGroup.setExclusive(True)  # 중복 선택 불가
        self.percentages = [0, 10]
        for percent in self.percentages:
            rb = QRadioButton(f"{percent}%")
            self.costProgressionGroup.addButton(rb)
            rb.toggled.connect(self.calculateRemuneration)
            costProgressionLayout.addWidget(rb)
        self.costProgressionGroup.buttons()[0].setChecked(True)  # 첫 번째 라디오 버튼을 디폴트로 선택
        costProgressionBox.setLayout(costProgressionLayout)
        mainLayout.addWidget(costProgressionBox)

    def setupNumPeople(self, mainLayout):
        self.numPeopleLayout = QHBoxLayout()
        self.numPeopleLabel = QLabel("인원수:")
        self.numPeopleSpinBox = QSpinBox()
        self.numPeopleSpinBox.setMinimum(2)
        self.numPeopleSpinBox.setValue(2)

        # QPalette를 사용하여 비활성화된 QSpinBox의 배경색과 텍스트 색상 설정
        palette = self.numPeopleSpinBox.palette()  # QSpinBox의 현재 팔레트 가져오기
        palette.setColor(QPalette.Disabled, QPalette.Base, Qt.gray)  # 배경색 설정
        palette.setColor(QPalette.Disabled, QPalette.Text, Qt.darkGray)  # 텍스트 색상 설정
        self.numPeopleSpinBox.setPalette(palette)

        self.numPeopleSpinBox.setEnabled(False)  # 처음에는 비활성화 상태로 설정
        self.numPeopleSpinBox.valueChanged.connect(self.calculateRemuneration)
        self.numPeopleLayout.addWidget(self.numPeopleLabel)
        self.numPeopleLayout.addWidget(self.numPeopleSpinBox)
        mainLayout.addLayout(self.numPeopleLayout)

    def setupDueDate(self, mainLayout):
        self.dueDateLabel = QLabel("납기일:")
        self.dueDateEdit = QDateEdit()
        self.dueDateEdit.setCalendarPopup(True)  # 캘린더 팝업 활성화
        self.dueDateEdit.setDate(QDate(datetime.now().year, 3, 31))  # 디폴트 날짜 설정
        self.dueDateEdit.dateChanged.connect(self.calculateRemuneration)  # 날짜 변경 시 계산 메소드 호출
        dueDateLayout = QHBoxLayout()
        dueDateLayout.addWidget(self.dueDateLabel)
        dueDateLayout.addWidget(self.dueDateEdit)
        mainLayout.addLayout(dueDateLayout)

    def setupBaseCalcProgression(self, mainLayout):
        baseCalcProgressionBox = QGroupBox("기본계산 누진 선택")
        baseCalcProgressionLayout = QHBoxLayout()
        self.baseCalcProgressionGroup = QButtonGroup(self)
        self.baseCalcProgressionGroup.setExclusive(False)  # 다중 선택 허용

        self.basePercentages = ['일반', '공동사업자(인원)', '건설업 등(40%)', '외감대상(50%)']
        for i, percent in enumerate(self.basePercentages):
            rb = QRadioButton(percent)
            self.baseCalcProgressionGroup.addButton(rb)
            baseCalcProgressionLayout.addWidget(rb)
            rb.toggled.connect(lambda checked, rb=rb: self.toggleBaseCalcProgression(rb, checked))

        baseCalcProgressionBox.setLayout(baseCalcProgressionLayout)
        mainLayout.addWidget(baseCalcProgressionBox)

    def setupOtherReductionsAndRemuneration(self, mainLayout):
        self.otherReductionsLayout = QVBoxLayout()

        # 기타감면 및 성실신고보수 활성화 버튼
        self.enableOtherReductionsButton = QRadioButton('기타감면 및 성실신고보수 추가')
        self.otherReductionsLayout.addWidget(self.enableOtherReductionsButton)

        # 기타감면 입력란
        self.otherReductionsEdit = QLineEdit()
        self.otherReductionsEdit.setPlaceholderText('기타감면')
        self.otherReductionsEdit.setVisible(False)  # 초기에는 숨김

        # 성실신고보수 입력란
        self.faithfulReportingRemunerationEdit = QLineEdit()
        self.faithfulReportingRemunerationEdit.setPlaceholderText('성실신고보수')
        self.faithfulReportingRemunerationEdit.setVisible(False)  # 초기에는 숨김

        self.otherReductionsLayout.addWidget(self.otherReductionsEdit)
        self.otherReductionsLayout.addWidget(self.faithfulReportingRemunerationEdit)

        # 라디오 버튼 상태 변경에 따른 입력란 보이기/숨기기
        self.enableOtherReductionsButton.toggled.connect(self.toggleOtherReductionsAndRemunerationFields)

        mainLayout.addLayout(self.otherReductionsLayout)

    def toggleOtherReductionsAndRemunerationFields(self, checked):
        self.otherReductionsEdit.setVisible(checked)
        self.faithfulReportingRemunerationEdit.setVisible(checked)
        if not checked:
            self.otherReductionsEdit.clear()
            self.faithfulReportingRemunerationEdit.clear()


    def toggleBaseCalcProgression(self, rb, checked):
        if rb.text() == "일반" and checked:
            for button in self.baseCalcProgressionGroup.buttons():
                if button != rb:
                    button.setChecked(False)
        elif checked:
            self.baseCalcProgressionGroup.buttons()[0].setChecked(False)

        # "일반" 외의 모든 버튼이 비활성화되어 있을 때 "일반" 버튼 자동 체크
        nonGeneralButtonsChecked = any(button.isChecked() for button in self.baseCalcProgressionGroup.buttons() if button.text() != "일반")
        if not nonGeneralButtonsChecked:
            self.baseCalcProgressionGroup.buttons()[0].setChecked(True)


        self.toggleNumPeopleSpinBox()
        self.calculateRemuneration()  # 추가 보수 라벨 업데이트를 위해 호출

    def toggleNumPeopleSpinBox(self):
        # "공동사업자(인원)" 버튼이 체크되어 있으면 인원수 입력란 활성화, 그렇지 않으면 비활성화
        isJointBusinessChecked = self.baseCalcProgressionGroup.buttons()[1].isChecked()
        self.numPeopleSpinBox.setEnabled(isJointBusinessChecked)

    def calculateRemuneration(self):
        if not hasattr(self, 'incomeEdit') or not self.incomeEdit.text():
            return

        if not self.validateInputs():
            return

        # 기본 설정
        entityType = '법인' if self.corporateButton.isChecked() else '개인'
        calculationType = 'B유형' if self.typeBButton.isChecked() else 'A유형'

        income = float(self.incomeEdit.text())
        remuneration = self.calculateBasicRemuneration(income, entityType, calculationType)

        additionalRemuneration = 0
        additionalDetails = ''

        # 추가 보수 계산
        if self.baseCalcProgressionGroup.buttons()[1].isChecked():  # 공동사업자(인원)
            numPeople = self.numPeopleSpinBox.value()
            jointBusinessAddition = remuneration * (numPeople - 1) * 0.2
            additionalRemuneration += jointBusinessAddition
            additionalDetails += f"공동사업자({numPeople}인): +{int(jointBusinessAddition)}\n"

        if self.baseCalcProgressionGroup.buttons()[2].isChecked():  # 건설업 등(40%)
            constructionAddition = remuneration * 0.4
            additionalRemuneration += constructionAddition
            additionalDetails += f"건설업 추가: +{int(constructionAddition)}\n"

        if self.baseCalcProgressionGroup.buttons()[3].isChecked():  # 외감대상(50%)
            auditAddition = remuneration * 0.5
            additionalRemuneration += auditAddition
            additionalDetails += f"외감대상 추가: +{int(auditAddition)}\n"

        # 기본 보수 및 최종 보수 계산 및 라벨 업데이트
        finalRemuneration = remuneration + additionalRemuneration
        details = f'기본보수: {int(remuneration)}\n'
        final_details = f"최종 기본 보수: {int(finalRemuneration)}"

        self.resultLabel.setText(details)

        # 추가 보수 라벨 업데이트
        if additionalRemuneration > 0:
            self.additionalLabel.setText(additionalDetails)
            self.additionalLabel.show()
        else:
            self.additionalLabel.hide()

        self.final_resultLabel.setText(final_details)
        self.finalRemuneration = finalRemuneration
        
        # 원가계산 누진 계산
        selectedCostButton = self.costProgressionGroup.checkedButton()
        if selectedCostButton:
            selectedPercentage = self.percentages[self.costProgressionGroup.buttons().index(selectedCostButton)]
            costProgression = int(finalRemuneration * selectedPercentage / 100)
            self.costProgressionLabel.setText(f'원가계산누진: {costProgression}')

    def calculateBasicRemuneration(self, income, entityType, calculationType):
        if calculationType == 'A유형':
            # A유형 개인
            if entityType == '개인':
                base_amount, rate, min = 300000, 0, 0
                if income >= 1e8:
                    base_amount, rate, min = 300000, 0.0015, 1e8
                if income >= 3e8:
                    base_amount, rate, min = 600000, 0.001, 3e8
                if income >= 5e8:
                    base_amount, rate, min = 800000, 0.0008, 5e8
                if income >= 10e8:
                    base_amount, rate, min = 1200000, 0.0006, 10e8
                if income >= 30e8:
                    base_amount, rate, min = 2400000, 0.0004, 30e8
                if income >= 50e8:
                    base_amount, rate, min = 3200000, 0.0002, 50e8
                if income >= 100e8:
                    base_amount, rate, min = 4200000, 0.00018, 100e8
                if income >= 500e8:
                    base_amount, rate, min = 11400000, 0.00016, 500e8
                if income >= 1000e8:
                    base_amount, rate, min = 19400000, 0.00014, 1000e8

            else:  # A유형 법인
                base_amount, rate, min = 400000, 0, 0
                if income >= 1e8:
                    base_amount, rate, min = 400000, 0.0015, 1e8
                if income >= 3e8:
                    base_amount, rate, min = 700000, 0.001, 3e8
                if income >= 5e8:
                    base_amount, rate, min = 900000, 0.0008, 5e8
                if income >= 10e8:
                    base_amount, rate, min = 1300000, 0.0006, 10e8
                if income >= 30e8:
                    base_amount, rate, min = 2500000, 0.0004, 30e8
                if income >= 50e8:
                    base_amount, rate, min = 3300000, 0.0002, 50e8
                if income >= 100e8:
                    base_amount, rate, min = 4300000, 0.00018, 100e8
                if income >= 500e8:
                    base_amount, rate, min = 11500000, 0.00016, 500e8
                if income >= 1000e8:
                    base_amount, rate, min = 19500000, 0.00014, 1000e8
        else:  # B유형 개인
            if entityType == '개인':
                base_amount, rate, min = 300000, 0, 0
                if income >= 1e8:
                    base_amount, rate, min = 300000, 0.002, 1e8
                if income >= 3e8:
                    base_amount, rate, min = 700000, 0.0016, 3e8
                if income >= 5e8:
                    base_amount, rate, min = 1020000, 0.0014, 5e8
                if income >= 10e8:
                    base_amount, rate, min = 1720000, 0.0012, 10e8
                if income >= 30e8:
                    base_amount, rate, min = 4120000, 0.001, 30e8
                if income >= 50e8:
                    base_amount, rate, min = 6120000, 0.0008, 50e8
                if income >= 100e8:
                    base_amount, rate, min = 10120000, 0.0005, 100e8
                if income >= 500e8:
                    base_amount, rate, min = 30120000, 0.0002, 500e8
                if income >= 1000e8:
                    base_amount, rate, min = 40120000, 0.00008, 1000e8

            else:  # B유형 법인
                base_amount, rate, min = 400000, 0, 0
                if income >= 1e8:
                    base_amount, rate, min = 400000, 0.002, 1e8
                if income >= 3e8:
                    base_amount, rate, min = 800000, 0.0016, 3e8
                if income >= 5e8:
                    base_amount, rate, min = 1120000, 0.0014, 5e8
                if income >= 10e8:
                    base_amount, rate, min = 1820000, 0.0012, 10e8
                if income >= 30e8:
                    base_amount, rate, min = 4220000, 0.001, 30e8
                if income >= 50e8:
                    base_amount, rate, min = 6220000, 0.0008, 50e8
                if income >= 100e8:
                    base_amount, rate, min = 10220000, 0.0005, 100e8
                if income >= 500e8:
                    base_amount, rate, min = 30220000, 0.0002, 500e8
                if income >= 1000e8:
                    base_amount, rate, min = 40220000, 0.00008, 1000e8
        
        if rate >= 0:
            excess_income = income - min
            remuneration = base_amount + excess_income * rate
        else:
            remuneration = base_amount

        remuneration = (remuneration // 10000) * 10000

        return remuneration
    
    
    def create_excel_file(self):

        # 설정 파일에서 설정 정보 로드
        try:
            with open('config.json', 'r') as configFile:
                config = json.load(configFile)
        except FileNotFoundError:
            self.messageLabel.setText("설정 파일을 찾을 수 없습니다. 회사 설정을 확인해주세요.")
            return

        if not self.validateInputs():
            return
        company_name = self.companyNameEdit.text()

        if not company_name:
            self.messageLabel.setText("거래처명을 입력해주세요.")
            return
        result = int(self.finalRemuneration)  # 소수점 제거
        costProgression = int(self.costProgressionLabel.text().split(': ')[1])  # 소수점 제거

        template_path = self.resource_path('양식.xlsx')
        output_path = f'{company_name}{datetime.now().year}년귀속 조정보수청구서.xlsx'

        # 파일이 이미 존재하는지 확인
        if os.path.exists(output_path):
            # 사용자에게 확인 받기
            reply = QMessageBox.question(self, '파일 덮어쓰기 확인', f"'{output_path}' 파일이 이미 존재합니다. 덮어쓰시겠습니까?",
                                        QMessageBox.Yes | QMessageBox.No, QMessageBox.No)

            if reply == QMessageBox.No:
                return  # 사용자가 '아니오'를 선택하면 작업 취소

        # 양식.xlsx 파일 로드
        workbook = load_workbook(filename=template_path)
        sheet = workbook.active

            # 설정 정보를 엑셀 파일에 할당
        sheet['A32'] = config.get('companyName', '')  # 회사명
        sheet['A2'] = config.get('address', '')  # 주소
        sheet['B4'] = config.get('documentNumber', '')  # 문서번호
        accountHolder = config.get('accountHolder', '')  # 예금주명
        sheet['C35'] = f"계좌번호(예금주:{accountHolder})"

    # 동적으로 추가된 계좌 정보 반영
        bankAccounts = config.get('bankAccounts', [])
        for i, account in enumerate(bankAccounts):
            row = 36 + i
            # 새로운 행을 추가하는 경우
            if row > 36:
                sheet.insert_rows(row)
                source_row_height = sheet.row_dimensions[row - 1].height
                sheet.row_dimensions[row].height = source_row_height


            # 병합할 범위의 첫 번째 셀에 스타일 복사
            for col in ['C', 'E']:  # 'C'열과 'E'열은 병합 범위의 시작점입니다.
                source_cell = sheet[f'{col}{row - 1}']
                target_cell = sheet[f'{col}{row}']
                copy_styles_from_cell(target_cell, source_cell)

            sheet[f'E{row}'].alignment = Alignment(horizontal='center', vertical='center')

            # 'C'열과 'D'열을 병합하고 'C'열에 값을 할당
            merge_range_cd = f'C{row}:D{row}'
            if not is_merged(sheet, merge_range_cd):
                sheet.merge_cells(merge_range_cd)
            sheet[f'C{row}'].value = account.get('bankName', '')

            # 'E'열과 'G'열을 병합하고 'E'열에 값을 할당
            merge_range_eg = f'E{row}:G{row}'
            if not is_merged(sheet, merge_range_eg):
                sheet.merge_cells(merge_range_eg)
            sheet[f'E{row}'].value = account.get('accountNumber', '')

            e_cell = sheet[f'E{row}']
            e_cell.font = Font(size=12)

            # 'B', 'F', 'G' 열에 대해서는 병합되지 않았으므로 스타일만 복사합니다.
            for col in ['B', 'F', 'G']:
                if col not in ['C', 'E']:  # 병합된 범위의 시작점을 제외
                    source_cell = sheet[f'{col}{row - 1}']
                    target_cell = sheet[f'{col}{row}']
                    copy_styles_from_cell(target_cell, source_cell)

        # 셀에 데이터 쓰기
        sheet['H4'] = datetime.now().date()
        sheet['B5'] = company_name
        sheet['G18'] = result
        sheet['G19'] = costProgression
        sheet['B7'] = f"{datetime.now().year-1}귀속 결산서 작성 및 세무조정 보수 청구 건"
        sheet['A9'] = f"1. 귀사의 무궁한 발전을 기원하며, {datetime.now().year-1}년 귀속 결산업무가 적정하게 종결될 수 있도록 협조하여"
        dueDate = self.dueDateEdit.date().toString("yyyy년 MM월 dd일")
        sheet['A12'] = f"2. {datetime.now().year-1}년 귀속 결산 및 세무조정에 대한 보수를 아래와 같이 청구하오니 {dueDate}까지"

        # 기타감면 및 성실신고보수 값 할당
        if self.enableOtherReductionsButton.isChecked():
            otherReductionsValue = self.otherReductionsEdit.text() or '0'
            faithfulReportingRemunerationValue = self.faithfulReportingRemunerationEdit.text() or '0'
            sheet['G24'] = int(otherReductionsValue)  # 기타감면 값
            sheet['G26'] = int(faithfulReportingRemunerationValue)  # 성실신고보수 값
        
        #파일저장
        try:
            workbook.save(filename=output_path)
            self.messageLabel.setText(f"'{output_path}' 파일 생성 완료.")
        except PermissionError:
            # 파일이 이미 열려 있는 경우
            self.messageLabel.setText(f"파일 '{output_path}'이(가)\n이미 열려 있습니다.\n 다른 이름으로 저장하세요.")
            # 사용자에게 다른 파일명 선택 요청
            new_output_path, _ = QFileDialog.getSaveFileName(self, "다른 이름으로 저장", output_path, "Excel Files (*.xlsx)")
            if new_output_path:  # 사용자가 새 파일명을 선택한 경우
                workbook.save(filename=new_output_path)
                self.messageLabel.setText(f"'{new_output_path}' 파일 생성 완료.")
            else:  # 사용자가 취소한 경우
                self.messageLabel.setText("파일 저장이 취소되었습니다.")

    def create_pdf_file(self):
            company_name = self.companyNameEdit.text().strip()
            if not company_name:
                self.messageLabel.setText("거래처명을 입력해주세요.")
                return
            
            if not self.validateInputs():
                return

            excel_file_path = f'{company_name}{datetime.now().year}년귀속 조정보수청구서.xlsx'
            pdf_file_path = f'{company_name}{datetime.now().year}년귀속 조정보수청구서.pdf'

            # 엑셀 파일이 존재하지 않으면 생성
            if not os.path.exists(excel_file_path):
                self.create_excel_file()

            # PDF 변환 시작 메시지
            self.messageLabel.setText("PDF 생성 중...")
    
            # Excel Application 시작
            excel = CreateObject("Excel.Application")
            excel.Visible = False

            try:
                # 엑셀 파일 열기
                wb = excel.Workbooks.Open(os.path.abspath(excel_file_path))

                # 엑셀 파일을 PDF로 저장
                wb.ExportAsFixedFormat(0, os.path.abspath(pdf_file_path))
                self.messageLabel.setText(f"'{pdf_file_path}' PDF 파일 생성 완료.")
            except Exception as e:
                self.messageLabel.setText(f"PDF 변환 중 오류 발생: {e}")
            finally:
                wb.Close(False)
                excel.Quit()

    def validateInputs(self):
        # 기타 감면 입력값 검증
        if self.enableOtherReductionsButton.isChecked():
            try:
                other_reductions = float(self.otherReductionsEdit.text())
            except ValueError:
                self.messageLabel.setText("기타 감면에 유효한 숫자를 입력해주세요.")
                return False

        # 성실신고보수 입력값 검증
        if self.enableOtherReductionsButton.isChecked():
            try:
                faithful_reporting = float(self.faithfulReportingRemunerationEdit.text())
                if faithful_reporting < 0:  # 성실신고보수 값이 음수인 경우
                    self.messageLabel.setText("성실신고보수에 유효한 양의 숫자를 입력해주세요.")
                    return False
            except ValueError:
                self.messageLabel.setText("성실신고보수에 유효한 숫자를 입력해주세요.")
                return False

        # 입력란에 이상한거 쓸 때
        try:
            float(self.incomeEdit.text())
        except ValueError:
            self.messageLabel.setText("기준금액에 유효한 숫자를 입력해주세요.")
            return False

        # 메세지라벨 리셋
        self.messageLabel.setText("")
        return True
    
def copy_styles_from_cell(new_cell, copy_cell):
    new_cell.font = copy(copy_cell.font)
    new_cell.border = copy(copy_cell.border)
    new_cell.fill = copy(copy_cell.fill)
    new_cell.number_format = copy_cell.number_format
    new_cell.protection = copy(copy_cell.protection)
    new_cell.alignment = copy(copy_cell.alignment)

def is_merged(sheet, cell_range):
    """주어진 셀 범위가 이미 병합되었는지 확인합니다."""
    for merged_range in sheet.merged_cells.ranges:
        if str(merged_range) == cell_range:
            return True
    return False

if __name__ == '__main__':
    app = QApplication(sys.argv)
    ex = Request_Money()
    ex.show()
    sys.exit(app.exec_())
